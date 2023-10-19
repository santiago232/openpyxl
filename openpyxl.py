import openpyxl

def agregar_gasto(worksheet, fecha, descripcion, monto):
    # Obtener la última fila ocupada
    ultima_fila = worksheet.max_row + 1

    # Llenar la fila con los detalles del gasto
    worksheet.cell(row=ultima_fila, column=1, value=fecha)
    worksheet.cell(row=ultima_fila, column=2, value=descripcion)
    worksheet.cell(row=ultima_fila, column=3, value=monto)

def calcular_resumen(worksheet):
    numero_total_gastos = worksheet.max_row - 1  # Restar 1 para excluir el encabezado
    monto_total_gastos = sum(worksheet.cell(row=row, column=3).value for row in range(2, worksheet.max_row + 1))

    gasto_mas_caro = max(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True), key=lambda row: row[2])
    gasto_mas_barato = min(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True), key=lambda row: row[2])

    return numero_total_gastos, gasto_mas_caro, gasto_mas_barato, monto_total_gastos

def main():
    # Crear o cargar el archivo Excel
    try:
        workbook = openpyxl.load_workbook("informe_gastos.xlsx")
        worksheet = workbook["Gastos"]
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Gastos"
        worksheet.append(["Fecha", "Descripción", "Monto"])

    while True:
        fecha = input("Ingrese la fecha del gasto (o 'q' para salir): ")
        if fecha.lower() == 'q':
            break

        descripcion = input("Ingrese la descripción del gasto: ")
        monto = float(input("Ingrese el monto del gasto: "))

        agregar_gasto(worksheet, fecha, descripcion, monto)

    # Calcular el resumen
    numero_total_gastos, gasto_mas_caro, gasto_mas_barato, monto_total_gastos = calcular_resumen(worksheet)

    print(f"\nResumen de gastos:")
    print(f"Número total de gastos: {numero_total_gastos}")
    print(f"Gasto más caro: Fecha - {gasto_mas_caro[0]}, Descripción - {gasto_mas_caro[1]}, Monto - {gasto_mas_caro[2]}")
    print(f"Gasto más barato: Fecha - {gasto_mas_barato[0]}, Descripción - {gasto_mas_barato[1]}, Monto - {gasto_mas_barato[2]}")
    print(f"Monto total de gastos: {monto_total_gastos}")

    # Guardar el archivo Excel
    workbook.save("informe_gastos.xlsx")
    print("Informe de gastos guardado en 'informe_gastos.xlsx'")

if _name_ == "_main_":
    main()
